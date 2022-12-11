from pathlib import Path

import glob
import numpy as np
import openpyxl
import pandas as pd

excel_path = f"{Path(__file__).parents[1]}/cash/input/CF表.xlsx"
# output_path = 'flaskr/cash/output/update_CF.xlsx'
# csv_path = "flaskr/cash/input/csv"

TEMPLATECOLUMNS = [
    '仕入れ日',
    'ASIN',
    '商品リンク',
    '商品名',
    '仕入れ額',
    ]

USECOLS = ['ASIN', 'タイトル', '仕入原価', '個数', '仕入日']

def write_to_default_excel(sheet, list_of_df, start_row, start_col):
    for y, row in enumerate(list_of_df):
        for x, cell in enumerate(row):
            if pd.isnull(cell) == True:
                pass
            elif pd.isnull(cell) != True:
                sheet.cell(row=start_row + y,
                        column=start_col + x,
                        value=list_of_df[y][x])


def read_excel_df():
    excel_df = pd.read_excel(excel_path, header=19, usecols=[
        '仕入れ日',
        'ASIN',
        '商品リンク',
        '商品名',
        '仕入れ額',
        ], dtype=str)
    excel_df['id'] = excel_df['仕入れ日'].str.cat(excel_df['ASIN'])
    return excel_df


def read_sedolist_csv(sedori_path: str):
    csv_file_list = glob.glob(sedori_path)
    sedolist_df = pd.read_csv(csv_file_list[0], encoding="cp932")
    sedolist_df = sedolist_df.rename(columns={'タイトル': '商品名', '仕入原価': '仕入れ額', '仕入日': '仕入れ日'})
    format_df = sedolist_df.loc[np.repeat(sedolist_df.index.values,sedolist_df['個数'])]
    format_df = format_df.reindex(columns=TEMPLATECOLUMNS)
    format_df['id'] = format_df['仕入れ日'].str.cat(format_df['ASIN'])
    return format_df


def select_write_spot(wc):
    for i in range(20,wc.max_row):
        if wc[f"K{i}"].value is None:           
            return i
        else:
            pass

def to_write(sedori_path: str, output_path: str):
    excel_df = read_excel_df()
    format_df = read_sedolist_csv(sedori_path)
    input_df = format_df[~format_df['id'].isin(excel_df['id'])]
    input_df['仕入れ日'] = pd.to_datetime(input_df['仕入れ日'])
    input_df = input_df.drop('id', axis=1)

    excel_file_list = glob.glob(excel_path)
    wb = openpyxl.load_workbook(excel_file_list[0])
    wc = wb.worksheets[0]
    value_list = input_df.values.tolist()

    idx = select_write_spot(wc)
    write_to_default_excel(wc, value_list, start_row=idx, start_col=11)
    wb.save(output_path)